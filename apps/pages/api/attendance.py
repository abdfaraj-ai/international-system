"""
attendance.py — API جهاز البصمة ZKTeco
مقيّد بدور M01 فقط
"""
import json
import re
import datetime
from datetime import date, timedelta
import pytz

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from ..models import ZKDevice, ZKEmployee, AttendanceRecord, EmployeeNote, ExcusedAbsence

LOCAL_TZ = pytz.timezone('Asia/Jerusalem')

# أحرف غير مقبولة في openpyxl
_ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]')

def _safe(val):
    """تنظيف القيمة قبل إدخالها في خلية Excel"""
    if val is None:
        return ''
    if not isinstance(val, str):
        return val
    # استبدل em-dash وen-dash بشرطة عادية
    val = val.replace('—', '-').replace('–', '-')
    return _ILLEGAL_CHARS.sub('', val)

def _fmt(dt):
    """تحويل datetime من UTC إلى توقيت القدس للعرض"""
    if dt is None:
        return None
    if timezone.is_aware(dt):
        dt = dt.astimezone(LOCAL_TZ)
    return dt.strftime('%Y-%m-%d %I:%M:%S %p')

def _fmt_date(dt):
    if dt is None:
        return None
    if timezone.is_aware(dt):
        dt = dt.astimezone(LOCAL_TZ)
    return dt.strftime('%Y-%m-%d')

def _fmt_time(dt):
    if dt is None:
        return None
    if timezone.is_aware(dt):
        dt = dt.astimezone(LOCAL_TZ)
    return dt.strftime('%I:%M %p')

import os
ZK_IP   = os.environ.get('ZK_IP',   '192.168.1.201')
ZK_PORT = int(os.environ.get('ZK_PORT', 4370))


def _require_m01(request):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'غير مصرح'}, status=401)
    if request.user.role != 'M01':
        return JsonResponse({'success': False, 'error': 'صلاحيات غير كافية'}, status=403)
    return None


def _get_or_create_device():
    device, _ = ZKDevice.objects.get_or_create(
        ip=ZK_IP,
        defaults={'name': 'الجهاز الرئيسي', 'port': ZK_PORT, 'password': 0}
    )
    return device


# ── مزامنة البيانات من الجهاز ─────────────────────────────────────────────────

@csrf_exempt
@require_http_methods(['POST'])
def api_attendance_sync(request):
    """POST /api/attendance/sync/ — جلب بيانات الجهاز وحفظها"""
    err = _require_m01(request)
    if err:
        return err

    try:
        from zk import ZK
    except ImportError:
        return JsonResponse({'success': False, 'error': 'مكتبة pyzk غير مثبتة'}, status=500)

    device = _get_or_create_device()
    zk     = ZK(device.ip, port=device.port, timeout=15, password=device.password)
    conn   = None

    try:
        conn = zk.connect()
        conn.disable_device()

        # ── مزامنة الموظفين مع قراءة الأسماء العربية الحقيقية ──
        import struct
        import zk.const as const

        pyzk_users = conn.get_users()
        raw_data, _ = conn.read_with_buffer(const.CMD_USERTEMP_RRQ, const.FCT_USER)
        RECORD_SIZE = 72

        # قراءة الأسماء من البيانات الخام بـ cp1256
        # هيكل السجل ZKTeco: bytes 0-3=uid (le uint32), byte 4=user_id, bytes 11-35=الاسم
        # نقرأ الاسم ونُطابقه مع uid من pyzk بطريقة الترتيب (record i → uid i+1)
        raw_names = {}
        pyzk_sorted = sorted(pyzk_users, key=lambda u: u.uid)
        for i in range(min(len(raw_data) // RECORD_SIZE, len(pyzk_sorted))):
            offset  = i * RECORD_SIZE
            rec_raw = raw_data[offset:offset + RECORD_SIZE]
            name_raw = rec_raw[11:35].lstrip(b'\x00').rstrip(b'\x00').strip()
            if not name_raw:
                continue
            try:
                name = name_raw.decode('cp1256').strip()
            except Exception:
                name = name_raw.decode('latin-1', errors='replace').strip()
            if name:
                raw_names[pyzk_sorted[i].uid] = name

        def _clean(s):
            """إزالة أحرف NUL وغيرها من الأحرف غير المقبولة في PostgreSQL"""
            if not s:
                return s
            return s.replace('\x00', '').strip()

        emp_count = 0
        for u in pyzk_users:
            real_name = _clean(raw_names.get(u.uid) or u.name or f'موظف-{u.uid}')
            _, created = ZKEmployee.objects.update_or_create(
                device=device, uid=u.uid,
                defaults={
                    'name':    real_name,
                    'user_id': _clean(str(u.user_id)),
                }
            )
            if created:
                emp_count += 1

        # ── مزامنة سجلات الحضور ──
        attendances = conn.get_attendance()
        att_new = 0
        emp_map = {e.user_id: e for e in ZKEmployee.objects.filter(device=device)}

        for a in attendances:
            uid_str  = _clean(str(a.user_id))
            emp_obj  = emp_map.get(uid_str)
            naive_ts = a.timestamp
            aware_ts = timezone.make_aware(naive_ts) if timezone.is_naive(naive_ts) else naive_ts

            _, created = AttendanceRecord.objects.get_or_create(
                device=device, user_id=uid_str, timestamp=aware_ts,
                defaults={
                    'employee': emp_obj,
                    'punch':    getattr(a, 'punch', 0),
                }
            )
            if created:
                att_new += 1

        device.last_sync = timezone.now()
        device.save(update_fields=['last_sync'])

        return JsonResponse({
            'success':       True,
            'new_employees': emp_count,
            'new_records':   att_new,
            'total_records': AttendanceRecord.objects.filter(device=device).count(),
            'last_sync':     _fmt(device.last_sync),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    finally:
        if conn:
            conn.enable_device()
            conn.disconnect()


# ── قائمة الموظفين ────────────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_employees(request):
    """GET /api/attendance/employees/ — قائمة الموظفين من الجهاز"""
    err = _require_m01(request)
    if err:
        return err

    device = _get_or_create_device()
    emps   = ZKEmployee.objects.filter(device=device).order_by('uid')

    data = []
    for e in emps:
        last = AttendanceRecord.objects.filter(
            device=device, user_id=e.user_id
        ).order_by('-timestamp').first()
        data.append({
            'id':        e.id,
            'uid':       e.uid,
            'name':      e.name,
            'user_id':   e.user_id,
            'last_seen': _fmt(last.timestamp) if last else None,
        })

    return JsonResponse({
        'success':   True,
        'employees': data,
        'last_sync': _fmt(device.last_sync) if device.last_sync else None,
    })


# ── سجلات الحضور ──────────────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_records(request):
    """
    GET /api/attendance/records/
    params: date_from, date_to, user_id, period (today|week|month)
    """
    err = _require_m01(request)
    if err:
        return err

    device = _get_or_create_device()
    qs     = AttendanceRecord.objects.filter(device=device).select_related('employee')

    # ── فلترة بالفترة ──
    period    = request.GET.get('period', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    user_id   = request.GET.get('user_id', '')

    today = timezone.now().astimezone(LOCAL_TZ).date()
    if period == 'today':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.min))
        end   = LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.max))
        qs = qs.filter(timestamp__gte=start, timestamp__lte=end)
    elif period == 'week':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today - timedelta(days=7), datetime.time.min))
        qs = qs.filter(timestamp__gte=start)
    elif period == 'month':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today - timedelta(days=30), datetime.time.min))
        qs = qs.filter(timestamp__gte=start)
    else:
        try:
            if date_from:
                qs = qs.filter(timestamp__gte=LOCAL_TZ.localize(datetime.datetime.combine(date.fromisoformat(date_from), datetime.time.min)))
            if date_to:
                qs = qs.filter(timestamp__lte=LOCAL_TZ.localize(datetime.datetime.combine(date.fromisoformat(date_to), datetime.time.max)))
        except ValueError:
            pass

    if user_id:
        qs = qs.filter(user_id=user_id)

    qs = qs.order_by('-timestamp')[:500]

    PUNCH_LABELS = {0: 'دخول', 1: 'خروج', 4: 'خروج استراحة', 5: 'دخول استراحة'}
    records = []
    for r in qs:
        records.append({
            'id':           r.id,
            'user_id':      r.user_id,
            'name':         r.employee.name if r.employee else r.user_id,
            'timestamp':    _fmt(r.timestamp),
            'date':         _fmt_date(r.timestamp),
            'time':         _fmt_time(r.timestamp),
            'punch':        r.punch,
            'punch_label':  PUNCH_LABELS.get(r.punch, str(r.punch)),
            'is_manual':    r.is_manual,
            'manual_reason': r.manual_reason,
        })

    return JsonResponse({'success': True, 'records': records, 'count': len(records)})


# ── تقرير يومي ────────────────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_report(request):
    """
    GET /api/attendance/report/
    params: date_from, date_to, period
    يُرجع ملخصاً بكل موظف: أول دخول، آخر خروج، عدد الأيام
    """
    err = _require_m01(request)
    if err:
        return err

    device = _get_or_create_device()
    qs     = AttendanceRecord.objects.filter(device=device).select_related('employee')

    period    = request.GET.get('period', 'week')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    today = timezone.now().astimezone(LOCAL_TZ).date()
    if period == 'today':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.min))
        end   = LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.max))
        qs = qs.filter(timestamp__gte=start, timestamp__lte=end)
    elif period == 'week':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today - timedelta(days=7), datetime.time.min))
        qs = qs.filter(timestamp__gte=start)
    elif period == 'month':
        start = LOCAL_TZ.localize(datetime.datetime.combine(today - timedelta(days=30), datetime.time.min))
        qs = qs.filter(timestamp__gte=start)
    else:
        try:
            if date_from:
                qs = qs.filter(timestamp__gte=LOCAL_TZ.localize(datetime.datetime.combine(date.fromisoformat(date_from), datetime.time.min)))
            if date_to:
                qs = qs.filter(timestamp__lte=LOCAL_TZ.localize(datetime.datetime.combine(date.fromisoformat(date_to), datetime.time.max)))
        except ValueError:
            pass

    # تجميع البيانات لكل موظف
    summary = {}
    for r in qs.order_by('timestamp'):
        uid = r.user_id
        if uid not in summary:
            summary[uid] = {
                'user_id':    uid,
                'name':       r.employee.name if r.employee else uid,
                'days':       set(),
                'first_in':   None,
                'last_out':   None,
                'total_punches': 0,
            }
        summary[uid]['days'].add(r.timestamp.date())
        summary[uid]['total_punches'] += 1
        if r.punch == 0:
            if not summary[uid]['first_in'] or r.timestamp < summary[uid]['first_in']:
                summary[uid]['first_in'] = r.timestamp
        if r.punch == 1:
            if not summary[uid]['last_out'] or r.timestamp > summary[uid]['last_out']:
                summary[uid]['last_out'] = r.timestamp

    report = []
    for uid, s in sorted(summary.items(), key=lambda x: x[1]['name']):
        report.append({
            'user_id':       s['user_id'],
            'name':          s['name'],
            'days_present':  len(s['days']),
            'first_in':      _fmt(s['first_in']) if s['first_in'] else '—',
            'last_out':      _fmt(s['last_out'])  if s['last_out']  else '—',
            'total_punches': s['total_punches'],
        })

    return JsonResponse({'success': True, 'report': report})


# ── حالة الجهاز ───────────────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_device_status(request):
    """GET /api/attendance/device-status/ — فحص الاتصال بالجهاز"""
    err = _require_m01(request)
    if err:
        return err

    try:
        from zk import ZK
        zk   = ZK(ZK_IP, port=ZK_PORT, timeout=5, password=0)
        conn = zk.connect()
        fw   = conn.get_firmware_version()
        t    = conn.get_time()
        conn.disconnect()
        return JsonResponse({
            'success':  True,
            'online':   True,
            'firmware': fw,
            'device_time': str(t),
            'ip': ZK_IP,
        })
    except Exception as e:
        return JsonResponse({
            'success': True,
            'online':  False,
            'error':   str(e),
            'ip':      ZK_IP,
        })


# ── إحصائيات الحضور والغياب والتأخير ─────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_stats(request):
    """
    GET /api/attendance/stats/
    params: period (today|week|month|custom), date_from, date_to
    يُرجع إحصائيات تفصيلية لكل موظف: حضور / غياب / تأخير / نسب
    """
    err = _require_m01(request)
    if err:
        return err

    device = _get_or_create_device()

    # ── تحديد نطاق التواريخ ──
    period    = request.GET.get('period', 'month')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    late_hour = int(request.GET.get('late_hour', 9))   # ساعة التأخير (افتراضي 9 صباحاً)
    late_min  = int(request.GET.get('late_min',  0))

    today = timezone.now().astimezone(LOCAL_TZ).date()

    if period == 'today':
        start_date = end_date = today
    elif period == 'week':
        start_date = today - timedelta(days=6)
        end_date   = today
    elif period == 'month':
        start_date = today - timedelta(days=29)
        end_date   = today
    else:
        try:
            if date_from and date_to:
                start_date = date.fromisoformat(date_from)
                end_date   = date.fromisoformat(date_to)
            else:
                start_date = today - timedelta(days=29)
                end_date   = today
        except ValueError:
            start_date = today - timedelta(days=29)
            end_date   = today

    # ── أيام العمل في النطاق (الأحد - الخميس فقط) ──
    work_days = []
    d = start_date
    while d <= end_date:
        if d.weekday() not in (4, 5):  # استثناء الجمعة=4 والسبت=5
            work_days.append(d)
        d += timedelta(days=1)
    total_work_days = len(work_days)

    # ── جلب كل السجلات في النطاق ──
    start_dt = LOCAL_TZ.localize(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt   = LOCAL_TZ.localize(datetime.datetime.combine(end_date,   datetime.time.max))
    qs = (AttendanceRecord.objects
          .filter(device=device, timestamp__gte=start_dt, timestamp__lte=end_dt)
          .select_related('employee')
          .order_by('timestamp'))

    # optional user_id filter for per-employee detail view
    filter_user_id = request.GET.get('user_id', '').strip()
    if filter_user_id:
        qs = qs.filter(user_id=filter_user_id)

    # ── تجميع البيانات لكل موظف ──
    LATE_TIME = datetime.time(late_hour, late_min)
    emp_data  = {}

    for r in qs:
        uid = r.user_id
        if uid not in emp_data:
            emp_data[uid] = {
                'name': r.employee.name.strip() if r.employee else uid,
                'days': {},   # date → {in_dt, out_dt, in_time, out_time}
            }
        local_dt = r.timestamp.astimezone(LOCAL_TZ)
        d = local_dt.date()
        if d not in emp_data[uid]['days']:
            emp_data[uid]['days'][d] = {'in_dt': None, 'out_dt': None, 'in_time': None, 'out_time': None}
        if r.punch == 0 and emp_data[uid]['days'][d]['in_dt'] is None:
            emp_data[uid]['days'][d]['in_dt']   = local_dt
            emp_data[uid]['days'][d]['in_time']  = local_dt.time()
        if r.punch == 1:
            emp_data[uid]['days'][d]['out_dt']  = local_dt
            emp_data[uid]['days'][d]['out_time'] = local_dt.time()

    def _secs_to_hm(secs):
        """تحويل ثوانٍ إلى نص 'Xس Yد'"""
        h = int(secs) // 3600
        m = (int(secs) % 3600) // 60
        return f"{h}س {m:02d}د"

    def _fmt_avg_time(times_secs):
        """متوسط قائمة ثوانٍ → نص وقت 12h"""
        if not times_secs:
            return '—'
        avg = sum(times_secs) // len(times_secs)
        h24 = avg // 3600
        mn  = (avg % 3600) // 60
        h12 = h24 % 12 or 12
        ap  = 'AM' if h24 < 12 else 'PM'
        return f"{h12:02d}:{mn:02d} {ap}"

    # ── بناء الإحصائيات ──
    stats = []
    total_hours_all = 0  # مجموع ساعات كل الموظفين للمتوسط العام

    for uid, data in sorted(emp_data.items(), key=lambda x: x[1]['name']):
        days_present = len(data['days'])
        late_days    = sum(1 for v in data['days'].values() if v['in_time'] and v['in_time'] > LATE_TIME)
        on_time_days = days_present - late_days

        # ── ساعات العمل اليومية ──
        daily_hours = []   # ثوانٍ لكل يوم عمل فيه
        for v in data['days'].values():
            if v['in_dt'] and v['out_dt'] and v['out_dt'] > v['in_dt']:
                diff = (v['out_dt'] - v['in_dt']).total_seconds()
                if diff <= 16 * 3600:   # تجاهل أي فارق أكبر من 16 ساعة (خطأ في البيانات)
                    daily_hours.append(diff)

        total_secs_period = sum(daily_hours)
        avg_daily_secs    = total_secs_period / len(daily_hours) if daily_hours else 0
        total_hours_all  += total_secs_period

        # متوسط وقت الدخول والخروج
        in_secs_list  = [v['in_time'].hour  * 3600 + v['in_time'].minute  * 60 for v in data['days'].values() if v['in_time']]
        out_secs_list = [v['out_time'].hour * 3600 + v['out_time'].minute * 60 for v in data['days'].values() if v['out_time']]

        days_present  = min(days_present, total_work_days)
        days_absent   = max(0, total_work_days - days_present)
        pct_present   = round(days_present / total_work_days * 100) if total_work_days else 0
        pct_absent    = round(days_absent  / total_work_days * 100) if total_work_days else 0
        pct_late      = round(late_days    / days_present    * 100) if days_present    else 0

        # ساعات الأسبوع والشهر مقدّرة من المتوسط اليومي
        est_weekly  = avg_daily_secs * 5   # 5 أيام عمل أسبوعياً
        est_monthly = avg_daily_secs * 22  # ~22 يوم عمل شهرياً

        stats.append({
            'user_id':            uid,
            'name':               data['name'],
            'days_present':       days_present,
            'days_absent':        days_absent,
            'late_days':          late_days,
            'on_time_days':       on_time_days,
            'pct_present':        pct_present,
            'pct_absent':         pct_absent,
            'pct_late':           pct_late,
            'avg_in':             _fmt_avg_time(in_secs_list),
            'avg_out':            _fmt_avg_time(out_secs_list),
            # ── ساعات العمل ──
            'total_hours':        _secs_to_hm(total_secs_period),
            'total_hours_raw':    round(total_secs_period / 3600, 1),
            'avg_daily_hours':    _secs_to_hm(avg_daily_secs),
            'avg_daily_raw':      round(avg_daily_secs / 3600, 1),
            'est_weekly_hours':   _secs_to_hm(est_weekly),
            'est_monthly_hours':  _secs_to_hm(est_monthly),
            'days_with_full_rec': len(daily_hours),   # أيام فيها دخول + خروج
        })

    # ── إحصائيات عامة ──
    all_emp_count    = ZKEmployee.objects.filter(device=device).count()
    present_today    = len(set(
        r.user_id for r in AttendanceRecord.objects.filter(
            device=device,
            timestamp__gte=LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.min)),
            timestamp__lte=LOCAL_TZ.localize(datetime.datetime.combine(today, datetime.time.max)),
        )
    ))

    return JsonResponse({
        'success':               True,
        'stats':                 stats,
        'period':                {'from': str(start_date), 'to': str(end_date)},
        'total_work_days':       total_work_days,
        'total_employees':       all_emp_count,
        'present_today':         present_today,
        'absent_today':          max(0, all_emp_count - present_today),
        'late_threshold':        f'{late_hour:02d}:{late_min:02d}',
        'total_hours_all':       _secs_to_hm(total_hours_all),
        'avg_hours_per_emp':     _secs_to_hm(total_hours_all / len(stats)) if stats else '—',
    })


# ── تصدير Excel ───────────────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_export_excel(request):
    """
    GET /api/attendance/export/excel/
    params: period, date_from, date_to, user_id, type (records|stats)
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io
    except ImportError:
        return JsonResponse({'success': False, 'error': 'مكتبة openpyxl غير مثبتة'}, status=500)

    export_type = request.GET.get('type', 'stats')
    period      = request.GET.get('period', 'month')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    user_id     = request.GET.get('user_id', '')
    late_hour   = int(request.GET.get('late_hour', 9))
    late_min    = int(request.GET.get('late_min',  0))

    device = _get_or_create_device()
    today  = timezone.now().astimezone(LOCAL_TZ).date()

    if period == 'today':
        start_date = end_date = today
    elif period == 'week':
        start_date = today - timedelta(days=6)
        end_date   = today
    elif period == 'month':
        start_date = today - timedelta(days=29)
        end_date   = today
    elif date_from and date_to:
        try:
            start_date = date.fromisoformat(date_from)
            end_date   = date.fromisoformat(date_to)
        except ValueError:
            start_date = today - timedelta(days=29)
            end_date   = today
    else:
        start_date = today - timedelta(days=29)
        end_date   = today

    start_dt = LOCAL_TZ.localize(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt   = LOCAL_TZ.localize(datetime.datetime.combine(end_date,   datetime.time.max))

    GOLD_FILL   = PatternFill('solid', fgColor='C9A84C')
    DARK_FILL   = PatternFill('solid', fgColor='0B1120')
    EVEN_FILL   = PatternFill('solid', fgColor='111827')
    ODD_FILL    = PatternFill('solid', fgColor='0D1628')
    HEADER_FONT = Font(name='Arial', bold=True, color='0B1120', size=11)
    TITLE_FONT  = Font(name='Arial', bold=True, color='C9A84C', size=14)
    BODY_FONT   = Font(name='Arial', size=10, color='E8EAF0')
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='C9A84C')
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active

    if export_type == 'stats':
        ws.title = 'إحصائيات الحضور'

        work_days = []
        d = start_date
        while d <= end_date:
            if d.weekday() not in (4, 5):
                work_days.append(d)
            d += timedelta(days=1)
        total_work_days = len(work_days)

        qs = (AttendanceRecord.objects
              .filter(device=device, timestamp__gte=start_dt, timestamp__lte=end_dt)
              .select_related('employee').order_by('timestamp'))
        if user_id:
            qs = qs.filter(user_id=user_id)

        LATE_TIME = datetime.time(late_hour, late_min)
        emp_data  = {}
        for r in qs:
            uid = r.user_id
            if uid not in emp_data:
                emp_data[uid] = {'name': r.employee.name.strip() if r.employee else uid, 'days': {}}
            local_dt = r.timestamp.astimezone(LOCAL_TZ)
            d = local_dt.date()
            if d not in emp_data[uid]['days']:
                emp_data[uid]['days'][d] = {'in': None, 'out': None, 'in_dt': None, 'out_dt': None}
            if r.punch == 0 and emp_data[uid]['days'][d]['in'] is None:
                emp_data[uid]['days'][d]['in']    = local_dt.time()
                emp_data[uid]['days'][d]['in_dt'] = local_dt
            if r.punch == 1:
                emp_data[uid]['days'][d]['out']    = local_dt.time()
                emp_data[uid]['days'][d]['out_dt'] = local_dt

        ws.merge_cells('A1:M1')
        ws['A1'] = _safe(f'تقرير الحضور والانصراف - من {start_date} الى {end_date}')
        ws['A1'].font = TITLE_FONT
        ws['A1'].fill = DARK_FILL
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells('A2:M2')
        ws['A2'] = _safe(f'أيام العمل: {total_work_days} يوم  |  حد التأخير: {late_hour:02d}:{late_min:02d}')
        ws['A2'].font = Font(name='Arial', color='9AA3B8', size=10)
        ws['A2'].fill = DARK_FILL
        ws['A2'].alignment = CENTER
        ws.row_dimensions[2].height = 18

        headers = ['#', 'اسم الموظف', 'أيام الحضور', 'أيام الغياب', 'أيام التأخير',
                   'نسبة الحضور %', 'نسبة الغياب %', 'متوسط الدخول', 'متوسط الخروج',
                   'ساعات الفترة', 'متوسط يومي', 'تقدير أسبوعي', 'تقدير شهري']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = GOLD_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[3].height = 22

        def _xl_secs_to_hm(secs):
            h = int(secs) // 3600
            m = (int(secs) % 3600) // 60
            return f"{h}س {m:02d}د"

        def _xl_fmt_avg(times):
            if not times:
                return '—'
            s = sum(t.hour * 3600 + t.minute * 60 for t in times) // len(times)
            h12 = s // 3600 % 12 or 12
            ap = 'AM' if s // 3600 < 12 else 'PM'
            return f"{h12:02d}:{(s % 3600) // 60:02d} {ap}"

        for i, (uid, data) in enumerate(sorted(emp_data.items(), key=lambda x: x[1]['name']), 1):
            days_present = min(len(data['days']), total_work_days)
            days_absent  = max(0, total_work_days - days_present)
            late_days    = sum(1 for v in data['days'].values() if v['in'] and v['in'] > LATE_TIME)
            pct_present  = round(days_present / total_work_days * 100) if total_work_days else 0
            pct_absent   = round(days_absent  / total_work_days * 100) if total_work_days else 0

            avg_in  = _xl_fmt_avg([v['in']  for v in data['days'].values() if v['in']])
            avg_out = _xl_fmt_avg([v['out'] for v in data['days'].values() if v['out']])

            # ساعات العمل
            daily_secs = []
            for v in data['days'].values():
                if v.get('in_dt') and v.get('out_dt') and v['out_dt'] > v['in_dt']:
                    diff = (v['out_dt'] - v['in_dt']).total_seconds()
                    if diff <= 16 * 3600:
                        daily_secs.append(diff)

            total_secs   = sum(daily_secs)
            avg_day_secs = total_secs / len(daily_secs) if daily_secs else 0
            tot_hours    = _xl_secs_to_hm(total_secs)   if daily_secs else '—'
            avg_day      = _xl_secs_to_hm(avg_day_secs) if daily_secs else '—'
            est_week     = _xl_secs_to_hm(avg_day_secs * 5)  if daily_secs else '—'
            est_month    = _xl_secs_to_hm(avg_day_secs * 22) if daily_secs else '—'

            fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            for col, val in enumerate([i, data['name'], days_present, days_absent, late_days,
                                        f'{pct_present}%', f'{pct_absent}%', avg_in, avg_out,
                                        tot_hours, avg_day, est_week, est_month], 1):
                cell = ws.cell(row=i + 3, column=col, value=_safe(val))
                cell.font = BODY_FONT
                cell.fill = fill
                cell.alignment = CENTER
                cell.border = BORDER

        for i, w in enumerate([5, 25, 14, 14, 14, 14, 14, 16, 16, 14, 14, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        filename = f'attendance_stats_{start_date}_{end_date}.xlsx'

    else:
        ws.title = 'سجلات الحضور'

        qs = (AttendanceRecord.objects
              .filter(device=device, timestamp__gte=start_dt, timestamp__lte=end_dt)
              .select_related('employee').order_by('-timestamp'))
        if user_id:
            qs = qs.filter(user_id=user_id)

        ws.merge_cells('A1:F1')
        ws['A1'] = _safe(f'سجلات الحضور والانصراف - من {start_date} الى {end_date}')
        ws['A1'].font = TITLE_FONT
        ws['A1'].fill = DARK_FILL
        ws['A1'].alignment = CENTER
        ws.row_dimensions[1].height = 32

        headers = ['#', 'اسم الموظف', 'رقم الموظف', 'التاريخ', 'الوقت', 'النوع']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = GOLD_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[2].height = 22

        PUNCH_LABELS = {0: 'دخول', 1: 'خروج', 4: 'خروج استراحة', 5: 'دخول استراحة'}
        for i, r in enumerate(qs[:2000], 1):
            fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            for col, val in enumerate([
                i,
                r.employee.name if r.employee else r.user_id,
                r.user_id,
                _fmt_date(r.timestamp),
                _fmt_time(r.timestamp),
                PUNCH_LABELS.get(r.punch, str(r.punch)),
            ], 1):
                cell = ws.cell(row=i + 2, column=col, value=_safe(val))
                cell.font = BODY_FONT
                cell.fill = fill
                cell.alignment = CENTER
                cell.border = BORDER

        for i, w in enumerate([5, 25, 14, 14, 14, 16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        filename = f'attendance_records_{start_date}_{end_date}.xlsx'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from urllib.parse import quote
    safe_filename = ''.join(c if c.isalnum() or c in '-_.' else '_' for c in filename)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{safe_filename}"; '
        f"filename*=UTF-8''{quote(filename)}"
    )
    return response


# ── تقرير الدوام اليومي السريع ────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_daily(request):
    """
    GET /api/attendance/daily/?date=YYYY-MM-DD
    يُرجع لكل موظف حالته اليوم: حضر / غائب / متأخر
    """
    err = _require_m01(request)
    if err:
        return err

    device = _get_or_create_device()

    late_hour = int(request.GET.get('late_hour', 9))
    late_min  = int(request.GET.get('late_min',  0))
    date_str  = request.GET.get('date', '')

    today = timezone.now().astimezone(LOCAL_TZ).date()
    try:
        target_date = date.fromisoformat(date_str) if date_str else today
    except ValueError:
        target_date = today

    LATE_TIME = datetime.time(late_hour, late_min)

    start_dt = LOCAL_TZ.localize(datetime.datetime.combine(target_date, datetime.time.min))
    end_dt   = LOCAL_TZ.localize(datetime.datetime.combine(target_date, datetime.time.max))

    records = (AttendanceRecord.objects
               .filter(device=device, timestamp__gte=start_dt, timestamp__lte=end_dt)
               .select_related('employee')
               .order_by('timestamp'))

    excused_ids = set(
        ExcusedAbsence.objects
        .filter(date_from__lte=target_date, date_to__gte=target_date)
        .values_list('employee_id', flat=True)
    )

    day_map = {}
    for r in records:
        uid = r.user_id
        if uid not in day_map:
            emp_id = r.employee_id if r.employee else None
            day_map[uid] = {
                'emp_id':   emp_id,
                'user_id':  uid,
                'name':     r.employee.name if r.employee else uid,
                'in_time':  None,
                'out_time': None,
                'in_raw':   None,
                'excused':  emp_id in excused_ids,
            }
        local_dt = r.timestamp.astimezone(LOCAL_TZ)
        if r.punch == 0 and day_map[uid]['in_time'] is None:
            day_map[uid]['in_time'] = local_dt.strftime('%I:%M %p')
            day_map[uid]['in_raw']  = local_dt.time()
        if r.punch == 1:
            day_map[uid]['out_time'] = local_dt.strftime('%I:%M %p')

    all_emps = ZKEmployee.objects.filter(device=device).order_by('uid')
    result = []

    for emp in all_emps:
        uid = emp.user_id
        emp_excused = emp.id in excused_ids
        if uid in day_map:
            in_raw = day_map[uid].get('in_raw')
            status = 'late' if (in_raw and in_raw > LATE_TIME) else 'present'
        else:
            status = 'excused' if emp_excused else 'absent'

        result.append({
            'user_id':  uid,
            'emp_id':   emp.id,
            'name':     emp.name,
            'status':   status,
            'in_time':  day_map[uid]['in_time']  if uid in day_map else None,
            'out_time': day_map[uid]['out_time'] if uid in day_map else None,
            'excused':  emp_excused,
        })

    counts = {
        'present': sum(1 for x in result if x['status'] == 'present'),
        'late':    sum(1 for x in result if x['status'] == 'late'),
        'absent':  sum(1 for x in result if x['status'] == 'absent'),
        'excused': sum(1 for x in result if x['status'] == 'excused'),
        'total':   len(result),
    }

    return JsonResponse({
        'success':   True,
        'date':      str(target_date),
        'counts':    counts,
        'employees': result,
    })


# ── ملاحظات الموظفين ──────────────────────────────────────────────────────────

@require_http_methods(['GET', 'POST'])
def api_employee_notes(request, emp_id):
    """
    GET  /api/attendance/employees/<emp_id>/notes/
    POST /api/attendance/employees/<emp_id>/notes/
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        emp = ZKEmployee.objects.get(id=emp_id)
    except ZKEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموظف غير موجود'}, status=404)

    if request.method == 'GET':
        notes = EmployeeNote.objects.filter(employee=emp).select_related('created_by')
        data = [{
            'id':         n.id,
            'note_type':  n.note_type,
            'type_label': n.get_note_type_display(),
            'body':       n.body,
            'created_by': (n.created_by.get_full_name() or n.created_by.username) if n.created_by else 'النظام',
            'created_at': n.created_at.astimezone(LOCAL_TZ).strftime('%Y-%m-%d %I:%M %p'),
        } for n in notes]
        return JsonResponse({'success': True, 'notes': data, 'employee': emp.name})

    try:
        body_data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'بيانات غير صحيحة'}, status=400)

    body = (body_data.get('body') or '').strip()
    if not body:
        return JsonResponse({'success': False, 'error': 'نص الملاحظة مطلوب'}, status=400)

    note = EmployeeNote.objects.create(
        employee=emp,
        note_type=body_data.get('note_type', 'note'),
        body=body,
        created_by=request.user,
    )
    return JsonResponse({'success': True, 'id': note.id, 'message': 'تم حفظ الملاحظة'})


@require_http_methods(['DELETE'])
def api_employee_note_delete(request, note_id):
    """DELETE /api/attendance/notes/<note_id>/"""
    err = _require_m01(request)
    if err:
        return err
    try:
        EmployeeNote.objects.get(id=note_id).delete()
        return JsonResponse({'success': True})
    except EmployeeNote.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الملاحظة غير موجودة'}, status=404)


# ── أذونات الغياب ─────────────────────────────────────────────────────────────

@require_http_methods(['GET', 'POST'])
def api_excused_absences(request, emp_id):
    """
    GET  /api/attendance/employees/<emp_id>/excused/
    POST /api/attendance/employees/<emp_id>/excused/
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        emp = ZKEmployee.objects.get(id=emp_id)
    except ZKEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموظف غير موجود'}, status=404)

    if request.method == 'GET':
        absences = ExcusedAbsence.objects.filter(employee=emp).select_related('approved_by')
        data = [{
            'id':           a.id,
            'absence_type': a.absence_type,
            'type_label':   a.get_absence_type_display(),
            'date_from':    str(a.date_from),
            'date_to':      str(a.date_to),
            'days':         (a.date_to - a.date_from).days + 1,
            'reason':       a.reason,
            'approved_by':  (a.approved_by.get_full_name() or a.approved_by.username) if a.approved_by else 'النظام',
            'created_at':   a.created_at.astimezone(LOCAL_TZ).strftime('%Y-%m-%d'),
        } for a in absences]
        return JsonResponse({'success': True, 'absences': data, 'employee': emp.name})

    try:
        body_data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'بيانات غير صحيحة'}, status=400)

    try:
        date_from = date.fromisoformat(body_data.get('date_from', ''))
        date_to   = date.fromisoformat(body_data.get('date_to', ''))
    except ValueError:
        return JsonResponse({'success': False, 'error': 'التواريخ غير صحيحة'}, status=400)

    if date_to < date_from:
        return JsonResponse({'success': False, 'error': 'تاريخ الانتهاء قبل البداية'}, status=400)

    absence = ExcusedAbsence.objects.create(
        employee=emp,
        absence_type=body_data.get('absence_type', 'annual'),
        date_from=date_from,
        date_to=date_to,
        reason=(body_data.get('reason') or '').strip(),
        approved_by=request.user,
    )
    days = (date_to - date_from).days + 1
    return JsonResponse({'success': True, 'id': absence.id, 'days': days,
                         'message': f'تم تسجيل الإذن ({days} أيام)'})


@require_http_methods(['DELETE'])
def api_excused_absence_delete(request, absence_id):
    """DELETE /api/attendance/excused/<absence_id>/"""
    err = _require_m01(request)
    if err:
        return err
    try:
        ExcusedAbsence.objects.get(id=absence_id).delete()
        return JsonResponse({'success': True})
    except ExcusedAbsence.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الإذن غير موجود'}, status=404)


# ── تعديل / حذف سجل بصمة ─────────────────────────────────────────────────────

@require_http_methods(['PATCH', 'DELETE'])
def api_attendance_record_edit(request, record_id):
    """
    PATCH  /api/attendance/records/<id>/edit/  — تعديل وقت أو نوع بصمة
    DELETE /api/attendance/records/<id>/edit/  — حذف سجل بصمة
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        record = AttendanceRecord.objects.select_related('employee').get(id=record_id)
    except AttendanceRecord.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'السجل غير موجود'}, status=404)

    PUNCH_LABELS = {0: 'دخول', 1: 'خروج', 4: 'خروج استراحة', 5: 'دخول استراحة'}

    if request.method == 'DELETE':
        emp_name  = record.employee.name if record.employee else record.user_id
        old_time  = _fmt(record.timestamp)
        record.delete()
        try:
            from ..models import AuditLog
            AuditLog.log(
                action='other', request=request,
                actor=request.user.username,
                target=emp_name,
                detail=f'حذف سجل بصمة ({PUNCH_LABELS.get(record.punch, record.punch)}) كان في {old_time}',
            )
        except Exception:
            pass
        return JsonResponse({'success': True, 'message': 'تم حذف السجل'})

    # PATCH — تعديل
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'بيانات غير صحيحة'}, status=400)

    date_str  = (body.get('date') or '').strip()
    time_str  = (body.get('time') or '').strip()
    punch     = body.get('punch')
    reason    = (body.get('reason') or '').strip()

    if not date_str or not time_str or punch is None:
        return JsonResponse({'success': False, 'error': 'التاريخ والوقت والنوع مطلوبة'}, status=400)

    if punch not in (0, 1, 4, 5):
        return JsonResponse({'success': False, 'error': 'نوع البصمة غير صحيح'}, status=400)

    try:
        naive_dt = datetime.datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M')
    except ValueError:
        return JsonResponse({'success': False, 'error': 'صيغة التاريخ أو الوقت غير صحيحة'}, status=400)

    aware_dt = LOCAL_TZ.localize(naive_dt)

    # تحقق من عدم تعارض مع سجل آخر بنفس الوقت (غير هذا السجل نفسه)
    if AttendanceRecord.objects.filter(
        device=record.device, user_id=record.user_id, timestamp=aware_dt
    ).exclude(id=record_id).exists():
        return JsonResponse({'success': False, 'error': 'يوجد سجل بصمة بنفس الوقت بالفعل'}, status=409)

    old_time  = _fmt(record.timestamp)
    old_punch = record.punch
    emp_name  = record.employee.name if record.employee else record.user_id

    record.timestamp     = aware_dt
    record.punch         = punch
    record.is_manual     = True
    record.manual_reason = reason or f'تعديل يدوي من {old_time}'
    record.added_by      = request.user
    record.save(update_fields=['timestamp', 'punch', 'is_manual', 'manual_reason', 'added_by'])

    try:
        from ..models import AuditLog
        AuditLog.log(
            action='other', request=request,
            actor=request.user.username,
            target=emp_name,
            detail=(
                f'تعديل بصمة: {PUNCH_LABELS.get(old_punch, old_punch)} في {old_time} '
                f'→ {PUNCH_LABELS.get(punch, punch)} في {_fmt(aware_dt)}'
                + (f' — {reason}' if reason else '')
            ),
        )
    except Exception:
        pass

    return JsonResponse({
        'success':   True,
        'message':   f'تم تعديل البصمة بنجاح',
        'record': {
            'id':           record.id,
            'timestamp':    _fmt(record.timestamp),
            'date':         _fmt_date(record.timestamp),
            'time':         _fmt_time(record.timestamp),
            'punch':        record.punch,
            'punch_label':  PUNCH_LABELS.get(record.punch, str(record.punch)),
            'is_manual':    record.is_manual,
            'manual_reason': record.manual_reason,
        },
    })


# ── بصمة يدوية (للمدير) ───────────────────────────────────────────────────────

@require_http_methods(['POST'])
def api_attendance_manual_punch(request):
    """
    POST /api/attendance/manual-punch/
    body: { employee_id, punch (0=دخول|1=خروج), date, time, reason }
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'بيانات غير صحيحة'}, status=400)

    employee_id = body.get('employee_id')
    punch_type  = body.get('punch')
    date_str    = (body.get('date') or '').strip()
    time_str    = (body.get('time') or '').strip()
    reason      = (body.get('reason') or '').strip()

    if employee_id is None or punch_type is None or not date_str or not time_str:
        return JsonResponse({'success': False, 'error': 'جميع الحقول مطلوبة'}, status=400)

    if punch_type not in (0, 1):
        return JsonResponse({'success': False, 'error': 'نوع البصمة غير صحيح (0=دخول, 1=خروج)'}, status=400)

    try:
        emp = ZKEmployee.objects.get(id=employee_id)
    except ZKEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموظف غير موجود'}, status=404)

    try:
        naive_dt = datetime.datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M')
    except ValueError:
        return JsonResponse({'success': False, 'error': 'صيغة التاريخ أو الوقت غير صحيحة'}, status=400)

    aware_dt = LOCAL_TZ.localize(naive_dt)

    # تحقق من عدم وجود بصمة بنفس الوقت بالضبط
    if AttendanceRecord.objects.filter(device=emp.device, user_id=emp.user_id, timestamp=aware_dt).exists():
        return JsonResponse({'success': False, 'error': 'يوجد سجل بصمة بنفس الوقت بالفعل'}, status=409)

    PUNCH_LABELS = {0: 'دخول', 1: 'خروج'}
    record = AttendanceRecord.objects.create(
        device        = emp.device,
        employee      = emp,
        user_id       = emp.user_id,
        timestamp     = aware_dt,
        punch         = punch_type,
        is_manual     = True,
        manual_reason = reason,
        added_by      = request.user,
    )

    # تسجيل في AuditLog
    try:
        from ..models import AuditLog
        AuditLog.log(
            action='other',
            request=request,
            actor=request.user.username,
            target=emp.name,
            detail=f'بصمة يدوية ({PUNCH_LABELS[punch_type]}) في {_fmt(aware_dt)} — {reason}',
        )
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'id':      record.id,
        'message': f'تم إضافة بصمة {PUNCH_LABELS[punch_type]} يدوية للموظف {emp.name} في {_fmt(aware_dt)}',
    })


# ── تنبيهات الغياب المتكرر ────────────────────────────────────────────────────

@require_http_methods(['GET'])
def api_absence_alerts(request):
    """
    GET /api/attendance/alerts/
    params: month (YYYY-MM), threshold (int, افتراضي 3)
    """
    err = _require_m01(request)
    if err:
        return err

    import calendar
    device    = _get_or_create_device()
    threshold = int(request.GET.get('threshold', 3))
    month_str = request.GET.get('month', '')

    today = timezone.now().astimezone(LOCAL_TZ).date()
    if month_str:
        try:
            year, mon = [int(x) for x in month_str.split('-')]
        except Exception:
            year, mon = today.year, today.month
    else:
        year, mon = today.year, today.month

    _, last_day = calendar.monthrange(year, mon)
    month_start = date(year, mon, 1)
    month_end   = date(year, mon, last_day)
    work_days   = [
        date(year, mon, d) for d in range(1, last_day + 1)
        if date(year, mon, d).weekday() not in (4, 5) and date(year, mon, d) <= today
    ]
    total_work = len(work_days)

    start_dt = LOCAL_TZ.localize(datetime.datetime.combine(month_start, datetime.time.min))
    end_dt   = LOCAL_TZ.localize(datetime.datetime.combine(min(month_end, today), datetime.time.max))

    records = (AttendanceRecord.objects
               .filter(device=device, timestamp__gte=start_dt, timestamp__lte=end_dt, punch=0)
               .select_related('employee').order_by('timestamp'))

    emp_days = {}
    emp_db_ids = {}
    for r in records:
        uid = r.user_id
        local_dt = r.timestamp.astimezone(LOCAL_TZ)
        d = local_dt.date()
        if d.weekday() not in (4, 5):
            emp_days.setdefault(uid, set()).add(d)
        if uid not in emp_db_ids:
            emp_db_ids[uid] = r.employee_id

    excused_map = {}
    for ea in ExcusedAbsence.objects.filter(date_from__lte=month_end, date_to__gte=month_start).select_related('employee'):
        uid = ea.employee.user_id
        d = max(ea.date_from, month_start)
        while d <= min(ea.date_to, month_end) and d <= today:
            if d.weekday() not in (4, 5):
                excused_map.setdefault(uid, set()).add(d)
            d += timedelta(days=1)

    work_days_set = set(work_days)
    alerts, normal = [], []

    for emp in ZKEmployee.objects.filter(device=device):
        uid     = emp.user_id
        present = emp_days.get(uid, set())
        excused = excused_map.get(uid, set()) & work_days_set
        absent  = max(total_work - len(present) - len(excused - present), 0)
        entry = {
            'user_id':    uid,
            'emp_id':     emp.id,
            'name':       emp.name,
            'present':    len(present),
            'excused':    len(excused),
            'absent':     absent,
            'total_work': total_work,
            'alert':      absent >= threshold,
        }
        (alerts if entry['alert'] else normal).append(entry)

    alerts.sort(key=lambda x: -x['absent'])
    normal.sort(key=lambda x: x['name'])

    return JsonResponse({
        'success':   True,
        'month':     f'{year}-{mon:02d}',
        'threshold': threshold,
        'work_days': total_work,
        'alerts':    alerts,
        'normal':    normal,
    })


# ── تقرير موظف فردي (Excel كامل) ─────────────────────────────────────────────

@require_http_methods(['GET'])
def api_attendance_employee_report(request):
    """
    GET /api/attendance/export/employee/
    params: user_id (مطلوب), period (week|month|custom), date_from, date_to
    يُولّد ملف Excel شامل للموظف: ملخص + تفاصيل يومية + إحصائيات
    """
    err = _require_m01(request)
    if err:
        return err

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io
    except ImportError:
        return JsonResponse({'success': False, 'error': 'مكتبة openpyxl غير مثبتة'}, status=500)

    user_id   = request.GET.get('user_id', '').strip()
    period    = request.GET.get('period', 'month')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    late_hour = int(request.GET.get('late_hour', 9))
    late_min  = int(request.GET.get('late_min',  0))

    if not user_id:
        return JsonResponse({'success': False, 'error': 'user_id مطلوب'}, status=400)

    device = _get_or_create_device()
    today  = timezone.now().astimezone(LOCAL_TZ).date()

    # ── تحديد النطاق ──
    if period == 'week':
        start_date = today - timedelta(days=6)
        end_date   = today
    elif period == 'month':
        start_date = today.replace(day=1)
        end_date   = today
    elif date_from and date_to:
        try:
            start_date = date.fromisoformat(date_from)
            end_date   = date.fromisoformat(date_to)
        except ValueError:
            start_date = today.replace(day=1)
            end_date   = today
    else:
        start_date = today.replace(day=1)
        end_date   = today

    # ── بيانات الموظف ──
    try:
        emp = ZKEmployee.objects.get(device=device, user_id=user_id)
        emp_name = emp.name
    except ZKEmployee.DoesNotExist:
        emp_name = user_id

    start_dt = LOCAL_TZ.localize(datetime.datetime.combine(start_date, datetime.time.min))
    end_dt   = LOCAL_TZ.localize(datetime.datetime.combine(end_date,   datetime.time.max))

    qs = (AttendanceRecord.objects
          .filter(device=device, user_id=user_id, timestamp__gte=start_dt, timestamp__lte=end_dt)
          .order_by('timestamp'))

    # ── أيام العمل في النطاق ──
    work_days = []
    d = start_date
    while d <= end_date:
        if d.weekday() not in (4, 5):
            work_days.append(d)
        d += timedelta(days=1)
    total_work_days = len(work_days)

    # ── تجميع السجلات يومياً ──
    LATE_TIME   = datetime.time(late_hour, late_min)
    PUNCH_LABELS = {0: 'دخول', 1: 'خروج', 4: 'خروج استراحة', 5: 'دخول استراحة'}
    days_data   = {}   # date → {in_dt, out_dt, punches[], hours}

    for r in qs:
        local_dt = r.timestamp.astimezone(LOCAL_TZ)
        d = local_dt.date()
        if d not in days_data:
            days_data[d] = {'in_dt': None, 'out_dt': None, 'punches': [], 'hours': None}
        days_data[d]['punches'].append({
            'time':  local_dt.strftime('%I:%M %p'),
            'label': PUNCH_LABELS.get(r.punch, str(r.punch)),
            'punch': r.punch,
        })
        if r.punch == 0 and days_data[d]['in_dt'] is None:
            days_data[d]['in_dt'] = local_dt
        if r.punch == 1:
            days_data[d]['out_dt'] = local_dt

    # حساب ساعات كل يوم
    for v in days_data.values():
        if v['in_dt'] and v['out_dt'] and v['out_dt'] > v['in_dt']:
            diff = (v['out_dt'] - v['in_dt']).total_seconds()
            if diff <= 16 * 3600:
                v['hours'] = diff

    # ── إحصائيات ملخّصة ──
    days_present  = min(len(days_data), total_work_days)
    days_absent   = max(0, total_work_days - days_present)
    late_days     = sum(1 for v in days_data.values()
                        if v['in_dt'] and v['in_dt'].time() > LATE_TIME)
    daily_hours   = [v['hours'] for v in days_data.values() if v['hours']]
    total_secs    = sum(daily_hours)
    avg_day_secs  = total_secs / len(daily_hours) if daily_hours else 0
    pct_present   = round(days_present / total_work_days * 100) if total_work_days else 0
    pct_absent    = round(days_absent  / total_work_days * 100) if total_work_days else 0
    pct_late      = round(late_days    / days_present    * 100) if days_present else 0

    def secs_hm(s):
        return f"{int(s)//3600}س {(int(s)%3600)//60:02d}د"

    def fmt12(dt):
        if dt is None: return '—'
        h = dt.hour % 12 or 12
        return f"{h:02d}:{dt.minute:02d} {'AM' if dt.hour < 12 else 'PM'}"

    # ══ بناء Excel ══
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'ملخص'
    ws2 = wb.create_sheet('التفاصيل اليومية')
    ws3 = wb.create_sheet('كل البصمات')

    # ── أنماط ──
    GOLD_FILL   = PatternFill('solid', fgColor='C9A84C')
    DARK_FILL   = PatternFill('solid', fgColor='0B1120')
    DARK2_FILL  = PatternFill('solid', fgColor='111827')
    EVEN_FILL   = PatternFill('solid', fgColor='111827')
    ODD_FILL    = PatternFill('solid', fgColor='0D1628')
    GREEN_FILL  = PatternFill('solid', fgColor='052e16')
    RED_FILL    = PatternFill('solid', fgColor='1f0a0a')
    GOLD_FILL2  = PatternFill('solid', fgColor='1c1506')
    HEADER_FONT = Font(name='Arial', bold=True, color='0B1120', size=11)
    TITLE_FONT  = Font(name='Arial', bold=True, color='C9A84C', size=16)
    BOLD_GOLD   = Font(name='Arial', bold=True, color='C9A84C', size=12)
    BOLD_WHITE  = Font(name='Arial', bold=True, color='E8EAF0', size=11)
    BODY_FONT   = Font(name='Arial', size=10, color='E8EAF0')
    MUTED_FONT  = Font(name='Arial', size=10, color='9AA3B8')
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    RIGHT       = Alignment(horizontal='right',  vertical='center')
    thin        = Side(style='thin', color='C9A84C')
    thin_gray   = Side(style='thin', color='1e293b')
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)
    BORDER_G    = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # ════ ورقة 1: الملخص ════
    ws1.sheet_view.rightToLeft = True
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 22

    # شعار + عنوان
    ws1.merge_cells('A1:B1')
    ws1['A1'] = 'نظام انترناشونال الموحد'
    ws1['A1'].font = Font(name='Arial', bold=True, color='C9A84C', size=18)
    ws1['A1'].fill = DARK_FILL
    ws1['A1'].alignment = CENTER
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells('A2:B2')
    ws1['A2'] = _safe(f'تقرير الحضور والانصراف - {emp_name}')
    ws1['A2'].font = Font(name='Arial', bold=True, color='E8EAF0', size=13)
    ws1['A2'].fill = DARK_FILL
    ws1['A2'].alignment = CENTER
    ws1.row_dimensions[2].height = 26

    ws1.merge_cells('A3:B3')
    period_label = 'أسبوعي' if period == 'week' else 'شهري' if period == 'month' else 'مخصص'
    ws1['A3'] = _safe(f'الفترة: {start_date} - {end_date}  ({period_label})  |  تاريخ الإصدار: {today}')
    ws1['A3'].font = Font(name='Arial', color='475569', size=10)
    ws1['A3'].fill = DARK_FILL
    ws1['A3'].alignment = CENTER
    ws1.row_dimensions[3].height = 18

    ws1.row_dimensions[4].height = 10

    def summary_row(ws, row, label, value, val_color='E8EAF0', fill=None):
        c1 = ws.cell(row=row, column=1, value=_safe(label))
        c2 = ws.cell(row=row, column=2, value=_safe(value))
        c1.font = MUTED_FONT
        c2.font = Font(name='Arial', bold=True, color=val_color, size=11)
        c1.alignment = RIGHT
        c2.alignment = CENTER
        c1.border = BORDER_G
        c2.border = BORDER_G
        f = fill or DARK2_FILL
        c1.fill = f
        c2.fill = f
        ws.row_dimensions[row].height = 22

    summary_row(ws1, 5,  'اسم الموظف',          emp_name,                     'C9A84C')
    summary_row(ws1, 6,  'رقم الموظف',           user_id,                      '9AA3B8')
    summary_row(ws1, 7,  'الفترة',               f'{start_date} → {end_date}', '9AA3B8')
    summary_row(ws1, 8,  'أيام العمل في الفترة', total_work_days,               '3B82F6')
    summary_row(ws1, 9,  'أيام الحضور',          days_present,                  '00C97A', GREEN_FILL)
    summary_row(ws1, 10, 'أيام الغياب',          days_absent,                   'FF4D6A', RED_FILL)
    summary_row(ws1, 11, 'أيام التأخير',         late_days,                     'F0C96B', GOLD_FILL2)
    summary_row(ws1, 12, 'نسبة الحضور',          f'{pct_present}%',             '00C97A', GREEN_FILL)
    summary_row(ws1, 13, 'نسبة الغياب',          f'{pct_absent}%',              'FF4D6A', RED_FILL)
    summary_row(ws1, 14, 'نسبة التأخير',         f'{pct_late}%',                'F0C96B', GOLD_FILL2)
    summary_row(ws1, 15, 'إجمالي ساعات العمل',  secs_hm(total_secs) if daily_hours else '—', 'A78BFA')
    summary_row(ws1, 16, 'متوسط ساعات/يوم',     secs_hm(avg_day_secs) if daily_hours else '—', 'A78BFA')
    summary_row(ws1, 17, 'تقدير أسبوعي',        secs_hm(avg_day_secs*5)  if daily_hours else '—', '2DD4BF')
    summary_row(ws1, 18, 'تقدير شهري',          secs_hm(avg_day_secs*22) if daily_hours else '—', '2DD4BF')
    summary_row(ws1, 19, 'حد التأخير المعتمد',  f'{late_hour:02d}:{late_min:02d}', 'F0C96B')

    # تقييم
    if pct_present >= 90 and pct_late <= 10:
        rating, r_color = 'ممتاز ⭐', '00C97A'
    elif pct_present >= 75 and pct_late <= 25:
        rating, r_color = 'جيد ✓', '3B82F6'
    elif pct_present >= 60:
        rating, r_color = 'مقبول ⚠', 'F59E0B'
    else:
        rating, r_color = 'ضعيف ✗', 'FF4D6A'
    summary_row(ws1, 20, 'التقييم العام', rating, r_color)

    # ════ ورقة 2: التفاصيل اليومية ════
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells('A1:G1')
    ws2['A1'] = _safe(f'التفاصيل اليومية - {emp_name}  ({start_date} - {end_date})')
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].fill = DARK_FILL
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 32

    h2 = ['التاريخ', 'اليوم', 'وقت الدخول', 'وقت الخروج', 'ساعات العمل', 'التأخير', 'ملاحظة']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = GOLD_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws2.row_dimensions[2].height = 22

    DAY_NAMES = {0:'الاثنين',1:'الثلاثاء',2:'الأربعاء',3:'الخميس',4:'الجمعة',5:'السبت',6:'الأحد'}

    row = 3
    for wd in work_days:
        v       = days_data.get(wd)
        present = v is not None
        in_t    = fmt12(v['in_dt'])  if v else '—'
        out_t   = fmt12(v['out_dt']) if v else '—'
        hrs     = secs_hm(v['hours']) if v and v['hours'] else ('—' if not present else 'بدون خروج')
        is_late = present and v['in_dt'] and v['in_dt'].time() > LATE_TIME
        note    = 'متأخر' if is_late else ('حاضر' if present else 'غائب')
        note_color = 'F0C96B' if is_late else ('00C97A' if present else 'FF4D6A')

        fill = GREEN_FILL if present and not is_late else (GOLD_FILL2 if is_late else RED_FILL)

        vals = [str(wd), DAY_NAMES[wd.weekday()], in_t, out_t, hrs, 'نعم' if is_late else '—', note]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.fill = fill
            c.alignment = CENTER
            c.border = BORDER_G
            if col == 7:
                c.font = Font(name='Arial', bold=True, color=note_color, size=10)
            else:
                c.font = BODY_FONT
        ws2.row_dimensions[row].height = 20
        row += 1

    for i, w in enumerate([14, 12, 14, 14, 14, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ════ ورقة 3: كل البصمات ════
    ws3.sheet_view.rightToLeft = True

    ws3.merge_cells('A1:E1')
    ws3['A1'] = _safe(f'جميع البصمات - {emp_name}')
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].fill = DARK_FILL
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 32

    h3 = ['#', 'التاريخ', 'اليوم', 'الوقت', 'النوع']
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = GOLD_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws3.row_dimensions[2].height = 22

    all_punches = []
    for d_date, v in sorted(days_data.items()):
        for p in v['punches']:
            all_punches.append((d_date, p))

    PUNCH_COLOR = {0: '00C97A', 1: 'FF4D6A', 4: 'F59E0B', 5: '3B82F6'}
    for i, (d_date, p) in enumerate(all_punches, 1):
        fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
        pcolor = PUNCH_COLOR.get(
            next((pk for pk, pl in PUNCH_LABELS.items() if pl == p['label']), None), 'E8EAF0'
        )
        vals = [i, str(d_date), DAY_NAMES[d_date.weekday()], p['time'], p['label']]
        for col, val in enumerate(vals, 1):
            c = ws3.cell(row=i + 2, column=col, value=val)
            c.fill = fill
            c.alignment = CENTER
            c.border = BORDER_G
            if col == 5:
                c.font = Font(name='Arial', bold=True, color=pcolor, size=10)
            else:
                c.font = BODY_FONT

    for i, w in enumerate([5, 14, 12, 14, 18], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── إرسال الملف ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    period_str   = 'weekly' if period == 'week' else 'monthly' if period == 'month' else 'custom'
    safe_name    = ''.join(c if c.isalnum() or c in '-_' else '_' for c in emp_name)
    filename     = f'report_{safe_name}_{period_str}_{start_date}_{end_date}.xlsx'
    arabic_name  = f'تقرير_{emp_name}_{period_str}_{start_date}_{end_date}.xlsx'

    from urllib.parse import quote
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"; '
        f"filename*=UTF-8''{quote(arabic_name)}"
    )
    return response


# ── تغيير اسم موظف ────────────────────────────────────────────────────────────

@require_http_methods(['POST'])
def api_attendance_rename_employee(request, user_id):
    """POST /api/attendance/employees/<user_id>/rename/ — تغيير اسم الموظف"""
    err = _require_m01(request)
    if err:
        return err

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'بيانات غير صالحة'}, status=400)

    new_name = data.get('name', '').strip()
    if not new_name:
        return JsonResponse({'success': False, 'error': 'الاسم لا يمكن أن يكون فارغاً'}, status=400)
    if len(new_name) > 24:
        return JsonResponse({'success': False, 'error': 'الاسم طويل جداً (24 حرفاً كحد أقصى)'}, status=400)

    try:
        emp = ZKEmployee.objects.get(user_id=str(user_id))
        emp.name = new_name
        emp.save(update_fields=['name'])
        return JsonResponse({'success': True, 'name': new_name})
    except ZKEmployee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'الموظف غير موجود'}, status=404)
